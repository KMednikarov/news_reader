import json
import os
from datetime import datetime

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

import util.constants as cons
from model.news_data import NewsData
from model.scraper_query import ScraperQuery
from scrapers.base_scraper import BaseScraper
from scrapers.ft_scraper import FinancialTimesScraper
from scrapers.il_sole_scraper import IlSoleScraper
from scrapers.scmp_scraper import SouthChinaMorningPostScraper
from util.logger import Log

last_scrape_dates_file = cons.last_scrape_dates_file_path
companies_list_file = cons.companies_file_path
report_file = cons.report_file_path
log = Log('NewsReader')

scrapers = [IlSoleScraper(), FinancialTimesScraper(), SouthChinaMorningPostScraper()]
# scrapers = [SouthChinaMorningPostScraper()]


def main():
    log.info("----------------------------")
    log.info("Launching News Reader")
    log.info("----------------------------")
    try:
        last_scrape_dates = get_previous_dates(last_scrape_dates_file)
        companies = load_companies(companies_list_file)
        # companies = ['apple','amazon']
        news_list = get_articles(companies, scrapers, last_scrape_dates)
        save_results(news_list)

        log.info("----------------------------")
        log.info("Exiting News Reader")
        log.info("----------------------------")
    except Exception as e:
        log.exception(e)


def get_articles(companies, sources: [BaseScraper], last_scrape_dates):
    articles_list = []
    for source in sources:
        source_name = source.get_name()
        if source_name not in last_scrape_dates:
            last_scrape_dates[source_name] = {}

        source.get_logger().info('Started scraping')
        for company in companies:
            company_last_search_date = cons.MIN_DATE
            from_date = (datetime.now() - relativedelta(months=cons.LAST_MONTHS)).strftime(cons.SHORT_DATE_FORMAT)
            if company in last_scrape_dates[source_name]:
                company_last_search_date = datetime.strptime(last_scrape_dates[source_name][company], cons.DATE_FORMAT)
                from_date = (datetime
                             .strptime(last_scrape_dates[source_name][company], cons.DATE_FORMAT)
                             .strftime(cons.SHORT_DATE_FORMAT))

            query = ScraperQuery(keyword=company
                                 , from_date=from_date
                                 , last_scraped_date=company_last_search_date)
            news_data: NewsData
            news_data = source.scrape(query)
            if len(news_data.articles) <= 0:
                continue
            articles_list.append(news_data)
            last_scrape_dates[source_name][company] = news_data.articles[0].date if len(
                news_data.articles) > 0 else None

        source.close_driver()
        source.get_logger().info('Scraping completed')

    save_previous_dates(last_scrape_dates_file, last_scrape_dates)

    return articles_list


def get_previous_dates(file_path):
    if not os.path.exists('sources'):
        os.makedirs('sources')

    try:
        with open(file_path, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        # If the file does not exist, create it and return an empty dictionary
        with open(file_path, 'w') as file:
            json.dump({}, file)
        return {}


def save_previous_dates(file_path, data):
    out_file = open(file_path, "w")
    json.dump(data, out_file, indent=4)
    out_file.close()


def load_companies(file_path):
    if not os.path.exists(file_path):
        data = {'Companies list': []}
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)

    df = pd.read_excel(file_path, engine='openpyxl')
    if df.empty:
        input('No companies to scrape for! Please fill /sources/companies_list.xlsx !')
        raise Exception('No companies to scrape for! Please fill /sources/companies_list.xlsx')

    return df.iloc[:, 0].tolist()


def save_results(news_list: [NewsData]):
    data = []
    for news in news_list:
        for article in news.articles:
            news_data = {'Company': news.company, 'Source': news.source, 'Title': article.title, 'Link': article.link,
                         'Date': article.date}
            data.append(news_data)

    if len(data) == 0:
        warning_message = ('No new articles were found for these companies since last scrape!\nReport will not be '
                           'created ...')
        log.warning(warning_message)
        input('Press any key to continue...')
        return

    df = pd.DataFrame(data)
    workbook = Workbook()

    # Group data by company
    grouped_data = df.groupby('Company')

    # Iterate over each group (company) and write data to separate sheets
    for company, group in grouped_data:
        # Check if the sheet already exists
        if company in workbook.sheetnames:
            sheet = workbook[company[:31]]
        else:
            # If the sheet doesn't exist, create a new one
            sheet = workbook.create_sheet(title=company)

        header_row = ['Source', 'Title', 'Link', 'Date']
        sheet.append(header_row)

        for _, row in group.iterrows():
            sheet.append([row['Source'], row['Title'], row['Link'], row['Date']])

    if not os.path.exists('reports'):
        log.info("Report folder doesn't exist. Proceeding to create report folder.")
        os.makedirs('reports')

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    workbook_name = f'{report_file}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    workbook.save(workbook_name)
    log.info('News report generated - [{}]'.format(workbook_name))
    input('Press any key to continue...')


main()
